# -*- coding: utf-8 -*-
"""从 QUESTS.xlsx 生成 TKFQuestDef C# 代码块"""
import openpyxl

XLSX = r"G:\modmake\TKF_hideoutd and task systems\CUTarkovHideoutMod\QUESTS.xlsx"
OUT = r"G:\modmake\TKF_hideoutd and task systems\CUTarkovHideoutMod\QUESTS_generated.cs"

wb = openpyxl.load_workbook(XLSX)
ws = wb["任务线"]

def g(r, c):
    v = ws.cell(r, c).value
    return "" if v is None else str(v).strip()

rows = []
for r in range(2, ws.max_row + 1):
    quest_id = g(r, 1)
    if not quest_id or quest_id.startswith("#"):
        continue
    rows.append({
        "id": quest_id,
        "contact": g(r, 2),
        "main": g(r, 3),
        "name": g(r, 4),
        "desc": g(r, 5),
        "prereq": g(r, 6),
        "objectives": g(r, 7),
        "credits": g(r, 8),
        "rep": g(r, 9),
        "items": g(r, 10),
    })

lines = []
lines.append("// ===== 由 QUESTS.xlsx 自动生成，请勿手改 =====\n")
for i, q in enumerate(rows):
    lines.append("new TKFQuestDef")
    lines.append("{")
    lines.append(f"    Id = \"{q['id']}\",")
    lines.append(f"    ContactId = \"{q['contact']}\",")
    lines.append(f"    Main = {(q['main'] == '是')},")
    lines.append(f"    NameKey = \"{q['name']}\",")
    lines.append(f"    DescKey = \"{q['desc']}\",")
    # 前置任务
    prereqs = [p.strip() for p in q["prereq"].split(";") if p.strip()]
    if prereqs:
        lines.append("    PrerequisiteQuestIds = new List<string> { " + ", ".join(f'"{p}"' for p in prereqs) + " },")
    # 目标
    lines.append("    Objectives = new List<TKFObjectiveDef>")
    lines.append("    {")
    for obj in q["objectives"].split(";"):
        obj = obj.strip()
        if not obj:
            continue
        parts = obj.split(":")
        # 类型:参数:数量:描述Key
        otype = parts[0].strip() if len(parts) > 0 else ""
        target = parts[1].strip() if len(parts) > 1 else ""
        count = parts[2].strip() if len(parts) > 2 else "1"
        dkey = parts[3].strip() if len(parts) > 3 else ""
        if not dkey:
            dkey = q["desc"]  # 回退到任务描述Key
        lines.append("        new TKFObjectiveDef")
        lines.append("        {")
        lines.append(f"            Type = TKFObjectiveType.{otype},")
        if target:
            lines.append(f"            Target = \"{target}\",")
        lines.append(f"            RequiredCount = {count},")
        lines.append(f"            DescriptionKey = \"{dkey}\",")
        lines.append("        },")
    lines.append("    },")
    # 奖励
    lines.append("    Reward = new TKFReward")
    lines.append("    {")
    if q["credits"]:
        lines.append(f"        Credits = {q['credits']},")
    if q["rep"]:
        lines.append(f"        Reputation = {q['rep']},")
    items = [it.strip() for it in q["items"].split(";") if it.strip()]
    if items:
        lines.append("        Items = new List<TKFCostEntry>")
        lines.append("        {")
        for it in items:
            parts = it.split(":")
            iid = parts[0].strip()
            amt = parts[1].strip() if len(parts) > 1 else "1"
            lines.append(f"            new TKFCostEntry {{ ItemId = \"{iid}\", Amount = {amt} }},")
        lines.append("        },")
    lines.append("    },")
    if i != len(rows) - 1:
        lines.append("},")
        lines.append("")
    else:
        lines.append("}")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines) + "\n")
print("OK:", len(rows), "quests ->", OUT)
